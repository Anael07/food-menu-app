import sys

# Try to import required libraries
try:
    import pandas as pd
    print("Using pandas to read Excel file...")
    
    file_path = r'e:\Masters\STUDIES\Emancipatory Digital Transformation\food-menu-app\menu-data.xlsx'
    
    # Read Excel file and get sheet names
    xls = pd.ExcelFile(file_path)
    print("Available sheets:", xls.sheet_names)
    print()
    
    if 'menus' in xls.sheet_names:
        df = pd.read_excel(file_path, sheet_name='menus')
        print("="*100)
        print("MENUS TAB CONTENTS")
        print("="*100)
        print()
        print(df.to_string())
    else:
        print("'menus' sheet not found")
        
except ImportError as e:
    print(f"pandas not available: {e}")
    print("Trying openpyxl...")
    
    try:
        import openpyxl
        
        file_path = r'e:\Masters\STUDIES\Emancipatory Digital Transformation\food-menu-app\menu-data.xlsx'
        wb = openpyxl.load_workbook(file_path)
        
        print("Available sheets:", wb.sheetnames)
        print()
        
        if 'menus' in wb.sheetnames:
            ws = wb['menus']
            print("="*100)
            print("MENUS TAB CONTENTS")
            print("="*100)
            print()
            
            # Read all data
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                print(f"Row {idx}: {row}")
        else:
            print("'menus' sheet not found")
            
    except ImportError:
        print("openpyxl also not available")
        sys.exit(1)
